from random import randint

import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.enums import ContentType
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.filters import Command
import asyncio

from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, Message

API_TOKEN = '8035340049:AAE_9dXJqnoNYuoAFr_sGNX1X49Hh09gwIk'
bot = Bot(token=API_TOKEN)
dp = Dispatcher()

class Состояния(StatesGroup):
    меню_администратора = State()
    получение_сообщения_для_рассылки = State()
    подтверждение_рассылки = State()
    меню_совместной_торговли = State()
    обработка_валютных_пар = State()
    вверх_вниз = State()
    выбор_экспирации = State()
    обновление_канала = State()
    получение_id_PO = State()
    меню_новых_пользователей = State()
    подтверждение_нового_пользователя = State()
    причина_отклонения = State()
    написание_причины_отказа = State()
    да_нет_сообщение_отказа = State()
    рассылка_скриншота = State()
    получение_скриншота_для_рассылки = State()


keyboard_user_menu_on = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text='Сигналы: вкл')],
        [KeyboardButton(text='Справка')]
    ],
    resize_keyboard=True
)
keyboard_user_menu_off = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text='Сигналы: выкл')],
        [KeyboardButton(text='Справка')]
    ],
    resize_keyboard=True
)
keyboard_yes_no = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text='Да') ,KeyboardButton(text='Нет')],
        [KeyboardButton(text='Назад.')],
    ],
    resize_keyboard=True
)
#Клавиатуры Админа/трейдера
keyboard_trader_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Сделать рассылку."), KeyboardButton(text="Разослать скриншот.")],
        [KeyboardButton(text="Начать совместную торговлю.")],
        [KeyboardButton(text="Проверка новых пользователей.")],
        [KeyboardButton(text="Назад.")],

    ],
    resize_keyboard=True
)
keyboard_joint_trade = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Начало торговли.")],
        [KeyboardButton(text="Выбор валютной пары")],
        [KeyboardButton(text="Завершение торговли.")],
        [KeyboardButton(text="Назад.")],
    ],
    resize_keyboard=True
)
keyboard_currency_marks = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🇪🇺EUR\\AUD🇦🇺"), KeyboardButton(text="🇪🇺EUR\\CAD🇨🇦")],
        [KeyboardButton(text="🇪🇺EUR\\CHF🇨🇭"), KeyboardButton(text="🇪🇺EUR\\GBP🇬🇧")],
        [KeyboardButton(text="🇪🇺EUR\\JPY🇯🇵"), KeyboardButton(text="🇪🇺EUR\\USD🇺🇸")],
        [KeyboardButton(text="🇺🇸USD\\JPY🇯🇵"), KeyboardButton(text="🇺🇸USD\\CHF🇨🇭")],
        [KeyboardButton(text="🇺🇸USD\\CAD🇨🇦"), KeyboardButton(text="AUD\\USD🇺🇸")],
        [KeyboardButton(text="🇬🇧GBP\\USD🇺🇸"), KeyboardButton(text="")],
        [KeyboardButton(text="🇪🇺EUR\\USD🇺🇸 OTC"), KeyboardButton(text="🇦🇺EUR\\TRY🇹🇷 OTC")],
        [KeyboardButton(text="🇪🇺EUR\\RUB🇷🇺 OTC"), KeyboardButton(text="🇪🇺EUR\\NZD🇳🇿 OTC")],
        [KeyboardButton(text="🇪🇺EUR\\JPY🇯🇵 OTC"), KeyboardButton(text="🇦🇺EUR\\HUF🇭🇺 OTC")],
        [KeyboardButton(text="🇪🇺EUR\\GBP🇬🇧 OTC"), KeyboardButton(text="🇪🇺EUR\\CHF🇨🇭 OTC")],
        [KeyboardButton(text="🇺🇸USD\\VND🇻🇳 OTC"), KeyboardButton(text="🇺🇸USD\\THB🇹🇭 OTC")],
        [KeyboardButton(text="🇺🇸USD\\SGD🇨🇬 OTC"), KeyboardButton(text="🇺🇸USD\\RUB🇷🇺 OTC")],
        [KeyboardButton(text="🇺🇸USD\\PKR🇵🇰 OTC"), KeyboardButton(text="🇺🇸USD\\PHP🇵🇭 OTC")],
        [KeyboardButton(text="🇺🇸USD\\MYR🇲🇾 OTC"), KeyboardButton(text="🇺🇸USD\\MXN🇲🇽 OTC")],
        [KeyboardButton(text="🇺🇸USD\\JPY🇯🇵 OTC"), KeyboardButton(text="🇺🇸USD\\INR🇮🇳 OTC")],
        [KeyboardButton(text="🇺🇸USD\\IDR🇮🇩 OTC"), KeyboardButton(text="🇺🇸USD\\EGB🇪🇬 OTC")],
        [KeyboardButton(text="🇺🇸USD\\DZD🇩🇿 OTC"), KeyboardButton(text="🇺🇸USD\\COP🇨🇴 OTC")],
        [KeyboardButton(text="🇺🇸USD\\CNH🇨🇳 OTC"), KeyboardButton(text="🇺🇸USD\\CLP🇨🇱 OTC")],
        [KeyboardButton(text="🇺🇸USD\\CHF🇨🇭 OTC"), KeyboardButton(text="🇺🇸USD\\CAD🇨🇦 OTC")],
        [KeyboardButton(text="🇺🇸USD\\BRL🇧🇷 OTC"), KeyboardButton(text="🇺🇸USD\\BDT🇧🇩 OTC")],
        [KeyboardButton(text="🇺🇸USD\\ARS🇦🇷 OTC"), KeyboardButton(text="TND\\USD🇺🇸 OTC")],
        [KeyboardButton(text="🇳🇿NZD\\USD🇺🇸 OTC"), KeyboardButton(text="MAD\\USD🇺🇸 OTC")],
        [KeyboardButton(text="🇬🇧GBP\\USD🇺🇸 OTC"), KeyboardButton(text="🇪🇺EUR\\USD🇺🇸 OTC")],
        [KeyboardButton(text="🇦🇺AUD\\USD🇺🇸 OTC")],
        [KeyboardButton(text="Назад.")],
    ],
    resize_keyboard=True
)
keyboard_up_down = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text='Предупредить заранее.')],
        [KeyboardButton(text='Вверх↗️') ,KeyboardButton(text='Вниз⬇️')],
        [KeyboardButton(text='Перекрытие верх🛡↗️') ,KeyboardButton(text='Перекрытие вниз🛡⬇️')],
        [KeyboardButton(text='По тренду вверх📈') ,KeyboardButton(text='По тренду вниз📉')],
        [KeyboardButton(text='От уровня поддержки вверх📈️🔺') ,KeyboardButton(text='От уровня сопротивления вниз📉🔻')],
        [KeyboardButton(text='На пробой сопротивления вверх') ,KeyboardButton(text='На пробой поддержки вниз')],
        [KeyboardButton(text='Назад.')],
    ],
    resize_keyboard=True
)
keyboard_time_expiration = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text='1м') ,KeyboardButton(text='2м')],
        [KeyboardButton(text='3м') ,KeyboardButton(text='4м')],
        [KeyboardButton(text='5м') ,KeyboardButton(text='10+м')],
        [KeyboardButton(text='Назад.')],
    ],
    resize_keyboard=True
)
keyboard_new_users = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Получить пользователя на проверку.")],
        [KeyboardButton(text="Посмотреть статистику.")],
        [KeyboardButton(text='Назад.')],
    ],
    resize_keyboard=True
)
keyboard_accept_new_users = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Подтвердить."), KeyboardButton(text="Отклонить.")],
        [KeyboardButton(text='Назад.')],
    ],
    resize_keyboard=True
)
keyboard_deviations = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Не писать причину."), KeyboardButton(text="Написать причину.")],
        [KeyboardButton(text='Назад.')],
    ],
    resize_keyboard=True
)
keyboard_yes_no_message_otkaza = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Всё верно, отправить"), KeyboardButton(text="Переписать причину")],
    ],
    resize_keyboard=True
)
keyboard_skreen = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Отправить")],
        [KeyboardButton(text="Предпросмотр")],
        [KeyboardButton(text="Назад.")],
    ],
    resize_keyboard=True
)

async def проверка_подписки(user_id: int):
    admins_data = openpyxl.load_workbook('admins_data.xlsx')
    sheet_a_d = admins_data.active
    id_channel = sheet_a_d.cell(row=3, column=2).value
    try:
        статус = await bot.get_chat_member(chat_id=id_channel, user_id=user_id)
        return статус.status in ["member", "administrator", "creator"]
    except TelegramBadRequest:
        return False


#Обработка состояний
#Меню администратора
@dp.message(Состояния.меню_администратора)
async def обработка_меню_администратора(message: types.Message, state:FSMContext):
    текст_кнопки = message.text
    if текст_кнопки == "Сделать рассылку.":
        await message.answer('Введи сообщение для создания рассылки трейдерам')
        await state.set_state(Состояния.получение_сообщения_для_рассылки)
    elif текст_кнопки == "Разослать скриншот.":
        await message.answer('Скинь скриншот для рассылки')
        await state.set_state(Состояния.получение_скриншота_для_рассылки)
    elif текст_кнопки == "Начать совместную торговлю.":
        await message.answer("Перевод в меню совместной торговли", reply_markup=keyboard_joint_trade)
        await message.answer("Если нужно тут будет краткий гайд по меню.")
        await state.set_state(Состояния.меню_совместной_торговли)
    elif текст_кнопки == "Проверка новых пользователей.":
        await message.answer("Перевожу тебя в меню проверки пользователей.",reply_markup=keyboard_new_users)
        await state.set_state(Состояния.меню_новых_пользователей)
    elif текст_кнопки == "Назад.":
        await message.answer('Переход в меню пользователя')
        await state.clear()
        await обработка_start(message, state)

@dp.message(Состояния.получение_скриншота_для_рассылки)
async def обработка_получение_скриншота_для_рассылки(message: types.Message, state:FSMContext):

    if message.text == 'Отправить':
        data = await state.get_data()
        id_photo = data.get('id_photo')
        текст = data.get("текст")
        user_data = openpyxl.load_workbook('user_data.xlsx')
        sheet_u_d = user_data.active
        доставлено = 0
        for row in range(2, sheet_u_d.max_row + 1):
            try:
                доставлено += 1
                await bot.send_photo(chat_id=int(sheet_u_d.cell(row=row, column=1).value), photo=id_photo, caption=текст)
            except TelegramForbiddenError:
                доставлено += -1
        await message.answer(
            f'Скриншот был успешно доставлен {доставлено} из {sheet_u_d.max_row - 1} пользователей.')
        await state.clear()
        await message.answer('Активировано меню трейдера!', reply_markup=keyboard_trader_menu)
        await state.set_state(Состояния.меню_администратора)
    elif message.text == "Назад.":
        await state.clear()
        await message.answer('Активировано меню трейдера!', reply_markup=keyboard_trader_menu)
        await state.set_state(Состояния.меню_администратора)
    elif message.text == "Предпросмотр":
        data = await state.get_data()
        id_photo = data.get('id_photo')
        текст = data.get("текст")
        await bot.send_photo(chat_id=message.chat.id, photo=id_photo, caption=текст)
    elif message.content_type == ContentType.PHOTO:
        id_photo = message.photo[-1].file_id
        caption = message.caption
        if caption:
            текст = caption
            await state.update_data(текст=текст)
        await state.update_data(id_photo=id_photo)
        await message.answer(f'Если всё готово, нажми кнопку отправить.'
                             f'Если нужно добавить описание, просто напиши его', reply_markup=keyboard_skreen)
    elif message.content_type == ContentType.TEXT:
        текст = message.text
        await state.update_data(текст=текст)
        await message.answer(f'Если всё готово, нажми кнопку отправить.'
                             f'Если хочешь посмотреть что увидит пользователь нажми кнопку для предпросмотра.',reply_markup=keyboard_skreen)



@dp.message(Состояния.меню_новых_пользователей)
async def обработка_кнопок_меню_новых_пользователей(message: types.Message, state:FSMContext):
    new_users = openpyxl.load_workbook('new_users.xlsx')
    sheet_N_U = new_users.active
    текст_кнопки = message.text
    if текст_кнопки == "Назад.":
        await state.set_state(Состояния.меню_администратора)
        await message.answer('Активировано меню трейдера!', reply_markup=keyboard_trader_menu)
    elif текст_кнопки == "Получить пользователя на проверку.":
        if sheet_N_U.max_row > 1:
            for i in range(1, 100):
                row = randint(2, sheet_N_U.max_row + 1)
                if sheet_N_U.cell(row=row, column=1).value is not None:
                    await message.answer(f"Пользователь с ID{int(sheet_N_U.cell(row=row, column=1).value)}"
                                         f"\nПрислал ID Pocket Option - {int(sheet_N_U.cell(row=row, column=2).value)}")
                    await state.update_data(id_пользователя=int(sheet_N_U.cell(row=row, column=1).value),
                                            id_PO=int(sheet_N_U.cell(row=row, column=2).value),
                                            row=row)
                    await message.answer('Подтвердить пользователя?',reply_markup=keyboard_accept_new_users)
                    await state.set_state(Состояния.подтверждение_нового_пользователя)
                    break
            else:
                await message.answer('Не нашёл пользоватлей в Базе на регистрацию.')

        else:
            await message.answer('Все пользователи были проверены.')
    elif текст_кнопки == "Посмотреть статистику.":
        await message.answer(f'Пользователей ожидающих проверку = {sheet_N_U.max_row - 1} человек.')



@dp.message(Состояния.подтверждение_нового_пользователя)
async def обработка_подтверждение_нового_пользователя(message: types.Message, state:FSMContext):
    user_data = openpyxl.load_workbook('user_data.xlsx')
    sheet_user_data = user_data.active
    текст_кнопки = message.text
    data = await state.get_data()
    id_пользователя = data.get('id_пользователя')
    id_PO = data.get('id_PO')
    row = data.get('row')
    if текст_кнопки == "Назад.":
        await message.answer("Перевожу тебя в меню проверки пользователей.", reply_markup=keyboard_new_users)
        await state.set_state(Состояния.меню_новых_пользователей)
    elif текст_кнопки == "Подтвердить.":
        sheet_user_data.append([id_пользователя,"выкл",1,id_PO])
        user_data.save('user_data.xlsx')
        new_users = openpyxl.load_workbook('new_users.xlsx')
        sheet_N_U = new_users.active
        sheet_N_U.cell(row=row, column=1).value = None
        sheet_N_U.cell(row=row, column=2).value = None
        new_users.save('new_users.xlsx')
        await bot.send_message(chat_id=id_пользователя, text="Тебе открыт доступ к закрытому сообществу и сигналам!"
                                                             "\nКанал закрытого сообщества - ссылка тут")
        await message.answer("Пользователь подтверждён.", reply_markup=keyboard_new_users)
        await state.set_state(Состояния.меню_новых_пользователей)
    elif текст_кнопки == "Отклонить.":
        new_users = openpyxl.load_workbook('new_users.xlsx')
        sheet_N_U = new_users.active
        sheet_N_U.cell(row=row, column=1).value = None
        sheet_N_U.cell(row=row, column=2).value = None
        await bot.send_message(chat_id=id_пользователя, text="Регистрация отклонена.")
        new_users.save('new_users.xlsx')
        await message.answer('Пользователь откленён\n'
                             'Напишешь причину отклонения?', reply_markup=keyboard_deviations)
        await state.set_state(Состояния.причина_отклонения)

@dp.message(Состояния.причина_отклонения)
async def причина_отклонения(message: types.Message, state:FSMContext):
    текст_кнопки = message.text
    if текст_кнопки == "Назад.":
        await message.answer("Перевожу тебя в меню проверки пользователей.", reply_markup=keyboard_new_users)
        await state.set_state(Состояния.меню_новых_пользователей)
    elif текст_кнопки == "Не писать причину.":
        await message.answer("Перевожу тебя в меню проверки пользователей.", reply_markup=keyboard_new_users)
        await state.set_state(Состояния.меню_новых_пользователей)
    elif текст_кнопки == "Написать причину.":
        await message.answer('Напиши причину одним сообщением.')
        await state.set_state(Состояния.написание_причины_отказа)

@dp.message(Состояния.написание_причины_отказа)
async def отправка_причины_отказа(message: types.Message, state:FSMContext):
    текс_сообщения = message.text
    await state.update_data(сообщение=текс_сообщения)
    await message.answer(f'Проверь, всё верно?'
                         f'\n\n{message.text}', reply_markup=keyboard_yes_no_message_otkaza)
    await state.set_state(Состояния.да_нет_сообщение_отказа)

@dp.message(Состояния.да_нет_сообщение_отказа)
async def обработка_да_нет_сообщениие_отказа(message: types.Message, state:FSMContext):
    data = await state.get_data()
    id_пользователя = data.get('id_пользователя')
    сообщение = str(data.get('сообщение'))
    if message.text == "Всё верно, отправить":
        await bot.send_message(chat_id=id_пользователя, text=сообщение)
        await message.answer("Перевожу тебя в меню проверки пользователей.", reply_markup=keyboard_new_users)
        await state.set_state(Состояния.меню_новых_пользователей)
    elif message.text == "Переписать причину":
        await message.answer('Напиши причину одним сообщением.')
        await state.set_state(Состояния.написание_причины_отказа)


#Обработка рассылки
@dp.message(Состояния.получение_сообщения_для_рассылки)
async def обработка_текста_расслки(message: types.Message, state:FSMContext):
    await state.update_data(текст_сообщения=message.text)
    await message.answer(f'Текст рассылки\n{message.text}\nВсё верно?',reply_markup=keyboard_yes_no)
    await state.set_state(Состояния.подтверждение_рассылки)

@dp.message(Состояния.подтверждение_рассылки)
async def подтверждение_рассылки_сообщения(message: types.Message, state:FSMContext):
    if message.text == "Да":
        user_data = openpyxl.load_workbook('user_data.xlsx')
        sheet_u_d = user_data.active
        data = await state.get_data()
        текст_сообщения = data.get("текст_сообщения")
        доставлено = 0
        for row in range(2, sheet_u_d.max_row + 1):
            try:
                доставлено += 1
                await bot.send_message(chat_id=int(sheet_u_d.cell(row=row, column=1).value), text=текст_сообщения)
            except TelegramForbiddenError:
                доставлено += -1
        await message.answer(f'Сообщение \n{текст_сообщения}\n Было успешно доставлено {доставлено} из {sheet_u_d.max_row-1} пользователей.')
    elif message.text == "Нет":
        await message.answer('Введи отредактированое сообщение')
        await state.set_state(Состояния.получение_сообщения_для_рассылки)
    elif message.text == "Назад.":
        await message.answer("Вы переведены в главное меню", reply_markup=keyboard_trader_menu)
        await state.set_state(Состояния.меню_администратора)


#Обработка совместной торговли
@dp.message(Состояния.меню_совместной_торговли)
async def обработка_меню_совместной_торговли(message: types.Message, state:FSMContext):
    user_data = openpyxl.load_workbook('user_data.xlsx')
    sheet_u_d = user_data.active
    if message.text == "Начало торговли.":
        await message.answer("Оповещяю всех пользователей о начале торговли.")
        кол_во_успешных_отправок = 0
        for row in range(2, sheet_u_d.max_row + 1):
            try:
                if sheet_u_d.cell(row=row, column=2).value == 'вкл':
                    await bot.send_message(chat_id=int(sheet_u_d.cell(row=row,column=1).value),text='Приготовяся трейдер!\nСкоро будут сигналы!')
                    кол_во_успешных_отправок += 1
            except TelegramForbiddenError:
                кол_во_успешных_отправок -= 1
        await asyncio.sleep(1)
        await message.answer(f"Статистика оповещения\n"
                             f"Всего пользователей 👤 - {sheet_u_d.max_row-1}\n"
                             f"Успешно доставлено ✅ - {кол_во_успешных_отправок}")
    elif message.text == "Выбор валютной пары":
        await message.answer('Выбери валютную пару для создания сигнала.'
                             '\nУчти сигналы здесь отправляються мгновенно и без статистики!'
                             '\nТвои действия долнжны быть максимально чёткими', reply_markup=keyboard_currency_marks)
        await state.set_state(Состояния.обработка_валютных_пар)
    elif message.text == "Завершение торговли.":
        await message.answer("Оповещяю всех пользователей о завершении торговли.")
        кол_во_успешных_отправок = 0
        for row in range(2, sheet_u_d.max_row + 1):
            try:
                if sheet_u_d.cell(row=row, column=2).value == 'вкл':
                    await bot.send_message(chat_id=int(sheet_u_d.cell(row=row, column=1).value),
                                           text='Мы завершаем нашу торговлю! \nЖду твоего отзыва трейдер!')
                    кол_во_успешных_отправок += 1
            except TelegramForbiddenError:
                кол_во_успешных_отправок -= 1
        await asyncio.sleep(1)
        await message.answer(f"Статистика оповещения\n"
                             f"Всего пользователей 👤 - {sheet_u_d.max_row - 1}\n"
                             f"Успешно доставлено ✅ - {кол_во_успешных_отправок}")
    elif message.text == "Назад.":
        await message.answer('Меню трейдера', reply_markup=keyboard_trader_menu)
        await state.set_state(Состояния.меню_администратора)


#Обработка валютных пар обычного рынка
@dp.message(Состояния.обработка_валютных_пар)
async def обработка_нажатия_валютных_пар(message: types.Message, state:FSMContext):
    if message.text == 'Назад.':
        await state.set_state(Состояния.меню_совместной_торговли)
        await message.answer("Перевод в меню совместной торговли", reply_markup=keyboard_joint_trade)
    else:
        await state.update_data(пара=message.text)
        await state.set_state(Состояния.выбор_экспирации)
        await message.answer(f'Валютная пара {message.text}\n'
                             f'Какое будет предпологаемое время экспирации?', reply_markup=keyboard_time_expiration)



@dp.message(Состояния.выбор_экспирации)
async def выбор_времени_экспирации(message: types.Message, state:FSMContext):
    if message.text == 'Назад.':
        await state.set_state(Состояния.обработка_валютных_пар)
        await message.answer('Выбери валютную пару для создания сигнала.'
                             '\nУчти сигналы здесь отправляються мгновенно и без статистики!'
                             '\nТвои действия долнжны быть максимально чёткими', reply_markup=keyboard_currency_marks)
    else:
        await state.update_data(экспирация=message.text)
        data = await state.get_data()
        пара = data.get('пара')
        await state.set_state(Состояния.вверх_вниз)
        await message.answer(f'Валютная пара - {пара}'
                             f'\nВремя экспирации - {message.text}'
                             '\nВыбери направление движения!', reply_markup=keyboard_up_down)


@dp.message(Состояния.вверх_вниз)
async def отправка_сигнала_пользователям(message: types.Message, state:FSMContext):
    if message.text == "Назад.":
        await state.set_state(Состояния.обработка_валютных_пар)
        await message.answer("Возврат в меню валютных пар.", reply_markup=keyboard_currency_marks)
    elif message.text == "Предупредить заранее.":
        data = await state.get_data()
        пара = data.get("пара")
        экспирация = data.get("экспирация")
        user_data = openpyxl.load_workbook('user_data.xlsx')
        sheet_u_d = user_data.active
        for row in range(2, sheet_u_d.max_row + 1):
            try:
                if sheet_u_d.cell(row=row, column=2).value == 'вкл':
                    await bot.send_message(chat_id=int(sheet_u_d.cell(row=row, column=1).value),
                                       text=f'Приготовся!'
                                            f'\nВалютная пара - {пара}\n'
                                            f'На время - {экспирация}'
                                            )
            except TelegramForbiddenError:
                pass
    else:
        data = await state.get_data()
        пара = data.get("пара")
        экспирация = data.get("экспирация")
        user_data = openpyxl.load_workbook('user_data.xlsx')
        sheet_u_d = user_data.active
        for row in range(2, sheet_u_d.max_row + 1):
            try:
                if sheet_u_d.cell(row=row, column=2).value == 'вкл':
                    await bot.send_message(chat_id=int(sheet_u_d.cell(row=row, column=1).value),
                                            text=f'Сигнал!\n'
                                            f'Валютная пара - {пара}\n'
                                            f'На время - {экспирация}'
                                            f'\nВход - {message.text}')
            except TelegramForbiddenError:
                pass
        else:
            await state.set_state(Состояния.обработка_валютных_пар)
            await message.answer('Выбери валютную пару для создания сигнала.'
                                 '\nУчти сигналы здесь отправляються мгновенно и без статистики!'
                                 '\nТвои действия долнжны быть максимально чёткими',
                                 reply_markup=keyboard_currency_marks)



@dp.message(Состояния.получение_id_PO)
async def получение_id_PO(message: types.Message, state: FSMContext):
    try:
        id_покет_опшн = int(message.text)
        new_users = openpyxl.load_workbook('new_users.xlsx')
        sheet_N_U = new_users.active
        user_data = openpyxl.load_workbook('user_data.xlsx')
        sheet_U_D = user_data.active
        #Проверяем пользователя на обман по айди, а так же выдаём доступ к закрытой группе если пользователь не соврал
        for row in range(2, sheet_N_U.max_row + 1):
            if message.from_user.id == int(sheet_N_U.cell(row=row, column=1).value):
                await message.answer('Дождись проверки регистрации!')
                await state.clear()
                break
        else:
            for row in range(2, sheet_U_D.max_row + 1):
                id_PO_в_user_data = int(sheet_U_D.cell(row=row, column=4).value)
                if id_покет_опшн == id_PO_в_user_data:
                    id_пользователя = int(sheet_U_D.cell(row=row, column=1).value)
                    if message.from_user.id == id_пользователя:
                        await message.answer("Ты есть в нашей базе! ВОТ ССЫЛКА НА ЗАКРЫТУЮ ГРУППУ!"
                                             "После подписки используй /start для получения сигналов.")
                        await state.clear()
                        break
                    else:
                        await message.answer('Проверь правильность ID')
                        break
            else:
                sheet_N_U.append([message.from_user.id, id_покет_опшн])
                new_users.save('new_users.xlsx')
                await message.answer('Айди получено! Жди сообщения о результате проверки!')
                await state.clear()


    except ValueError:
        await message.answer('Введи корректный ID!')

#Обработка команд для всех
@dp.message(Command('start'))
async def обработка_start(message: types.Message, state: FSMContext):

    user_id=message.from_user.id
    подписан = await проверка_подписки(user_id)
    if подписан:
        user_data = openpyxl.load_workbook('user_data.xlsx')
        sheet_user_data = user_data.active
        for row in range(2, sheet_user_data.max_row + 1):
            id_в_табл = sheet_user_data.cell(row=row, column=1).value
            if message.from_user.id == id_в_табл:
                if int(sheet_user_data.cell(row=row, column=3).value) == 1:
                    sent_message = await message.answer('Авторизую тебя в боте, с возвращением трейдер!')
                    await asyncio.sleep(0.5)
                    await bot.delete_message(message.chat.id, sent_message.message_id)
                    if sheet_user_data.cell(row=row, column=2).value == 'вкл':
                        sent_message = await message.answer('Ты авторизован! Сигналы включены.')
                        await message.answer('Меню активировано.', reply_markup=keyboard_user_menu_on)
                        await asyncio.sleep(10)
                        await bot.delete_message(message.chat.id, sent_message.message_id)
                    else:
                        sent_message = await message.answer('Ты авторизован! Сигналы выключены.')
                        await message.answer('Меню активировано.', reply_markup=keyboard_user_menu_off)
                        await asyncio.sleep(10)
                        await bot.delete_message(message.chat.id, sent_message.message_id)
                    break
                else:
                    await message.answer('Ты заблокирован...')
                    break
        else:
            await message.answer('Вижу ты подписан на группу но тебя нет в базе!\n'
                                 'Введи айди своего аккаунта Pocket Option и после проверки тебе будет выдан доступ.')
            await state.set_state(Состояния.получение_id_PO)
    else:
        await message.answer("Приветствую👋🏻"
                            "\nЭтот бот выдает сделки до 90% проходимости."
                            "\nЧто бы продолжить использование:"
                            "\n1. Зарегистрироваться аккаунт Pocket Option по даной ссылке- https://clck.ru/3BKd8G (если ты уже зарегистрирован, приступай к следующему шагу)."
                            '\n2. Прислать ID своего аккаута(Без "ID" только цифры)'
                            "\n3. Дождаться проверки, и смело зарабатывать с нами!"
                             )
        await state.set_state(Состояния.получение_id_PO)


@dp.message(Command('help'))
async def обработка_help(message: types.Message):
    await message.answer('Сообщение с поддержкой')
    admins_data = openpyxl.load_workbook('admins_data.xlsx')
    sheet_a_d = admins_data.active
    if message.from_user.id == int(sheet_a_d.cell(row=1, column=2).value):
        await message.answer('Команда активации меню трейдера /start_admins_menu')


#Активация меню трейдера
@dp.message(Command('start_admins_menu'))
async def обработка_start_admins_menu(message: types.Message, state: FSMContext):
    admins_data = openpyxl.load_workbook('admins_data.xlsx')
    sheet_a_d = admins_data.active
    admin_id = int(sheet_a_d.cell(row=2, column=2).value)
    trader_id = int(sheet_a_d.cell(row=1, column=2).value)
    if message.from_user.id == admin_id or message.from_user.id == trader_id:
        await message.answer('Активировано меню трейдера!', reply_markup=keyboard_trader_menu)
        await state.set_state(Состояния.меню_администратора)



@dp.message(lambda message: message.text=='Назад.')
async def универсальное_назад(message: Message):
    await message.answer("Если бот заглючил или что-то не работает попробуй /start")
#Общий обработчки кнопок для пользователей
@dp.message()
async def обработчик_кнопок_пользователя(message: types.Message):
    подписан = await проверка_подписки(message.from_user.id)
    if подписан:
        текст_кнопки = message.text
        if текст_кнопки == "Справка":
            await message.answer("Привет! Данный бот это разработка для Алексея торговца от АМК")
        elif текст_кнопки == "Сигналы: выкл":
            user_data = openpyxl.load_workbook('user_data.xlsx')
            sheet_user_data = user_data.active
            for row in range(2, sheet_user_data.max_row + 1):
                if message.from_user.id == int(sheet_user_data.cell(row=row, column=1).value):
                    sheet_user_data.cell(row=row, column=2).value = 'вкл'
                    await message.answer('Получение сигналов активировано!', reply_markup=keyboard_user_menu_on)
                    user_data.save('user_data.xlsx')
                    break
        elif текст_кнопки == "Сигналы: вкл":
            user_data = openpyxl.load_workbook('user_data.xlsx')
            sheet_user_data = user_data.active
            for row in range(2, sheet_user_data.max_row + 1):
                if message.from_user.id == int(sheet_user_data.cell(row=row, column=1).value):
                    sheet_user_data.cell(row=row, column=2).value = 'выкл'
                    await message.answer('Получение сигналов отключено!', reply_markup=keyboard_user_menu_off)
                    user_data.save('user_data.xlsx')
                    break
        elif текст_кнопки == "Назад.":
            await message.answer("Если бот заглючил или что-то не работает попробуй /start")
    else:
        await message.answer('Ты не можешь пользоваться ботом!')


import os
import time
import socket
from threading import Thread

# Функция для имитации серверного порта
def simulate_server():
    # Настроим сокет для прослушивания порта
    server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server_socket.bind(('0.0.0.0', 8080))  # Открываем порт 8080
    server_socket.listen(1)  # Максимальное количество соединений

    print("Сервер запущен. Ожидание соединения...")

    while True:
        try:
            client_socket, client_address = server_socket.accept()  # Принять подключение
            print(f"Соединение с {client_address} установлено.")

            # Чтение данных от клиента (можно просто игнорировать)
            data = client_socket.recv(1024)
            print(f"Получено сообщение от клиента: {data.decode()}")

            # Ответ клиенту
            client_socket.send("Соединение успешно установлено!".encode('utf-8'))

            # Закрыть соединение
            client_socket.close()
        except Exception as e:
            print(f"Ошибка сервера: {e}")
            time.sleep(5)  # Подождать немного, если возникла ошибка

# Создаем и запускаем поток для имитации сервера
def start_simulation():
    server_thread = Thread(target=simulate_server)
    server_thread.daemon = True  # Завершается, когда программа завершена
    server_thread.start()

if __name__ == "__main__":
    start_simulation()  # Запускаем сервер в фоновом режиме
    # Далее идет код для запуска твоего бота
    bot.polling(none_stop=True)  # Важно: bot.polling должен быть в конце


print('Бот запущен!')
async def main():
    await dp.start_polling(bot)
if __name__ == '__main__':
    asyncio.run(main())
